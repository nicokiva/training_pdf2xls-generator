"""
helpers/xlsx.py — Writing data to the local .xlsx file.

Responsibilities:
    - Create or replace a tab in the workbook with the training data
"""

import openpyxl
from helpers.sheets import build_sheet_values


def write_xlsx_tab(wb, tab_name, days_data):
    """
    Creates or replaces a tab in the .xlsx workbook with the training data.
    Structure: day header → set row → Rep./Peso row → exercise rows.

    Parameters:
        wb        — openpyxl workbook (the open .xlsx file)
        tab_name  — name of the tab to create (e.g. "25-05-26-...")
        days_data — dict {day: [exercises]} as returned by parse_pdf
    """
    # Remove the blank default sheet from a fresh workbook so the output only
    # contains the generated routine tab.
    if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
        default = wb[wb.sheetnames[0]]
        if default.max_row == 1 and default.max_column == 1 and default["A1"].value is None:
            del wb[wb.sheetnames[0]]

    # Sheet names in XLSX cannot contain "/"
    xlsx_tab_name = tab_name.replace("/", "-")
    if xlsx_tab_name in wb.sheetnames:
        del wb[xlsx_tab_name]   # delete if already exists
    ws = wb.create_sheet(xlsx_tab_name, 0)   # 0 = insert at the front

    for row_idx, row_values in enumerate(build_sheet_values(days_data), start=1):
        for col_idx, val in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    return ws
