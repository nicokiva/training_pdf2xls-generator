"""
Tests for helpers/xlsx/xlsx.py — mocks openpyxl to avoid real file I/O.
"""
import pytest
from unittest.mock import MagicMock, patch


def _make_days_data():
    return {
        1: [
            {"name": "Sentadilla", "is_comb": False, "week_reps": [[10, 10, 8], [9, 9, 8], None, None]},
            {"name": "Press plano", "is_comb": False, "week_reps": [[8, 8, 8], None, None, None]},
        ],
        2: [
            {"name": "Dominada", "is_comb": False, "week_reps": [[6, 6, 5], None, None, None]},
        ]
    }


class TestWriteXlsxTab:
    def test_creates_sheet_in_workbook(self):
        from helpers.xlsx import write_xlsx_tab
        import openpyxl
        wb = openpyxl.Workbook()
        ws = write_xlsx_tab(wb, "19/05/26-...", _make_days_data())
        # Tab name replaces "/" with "-"
        assert "19-05-26-..." in wb.sheetnames

    def test_returns_worksheet(self):
        from helpers.xlsx import write_xlsx_tab
        import openpyxl
        wb = openpyxl.Workbook()
        ws = write_xlsx_tab(wb, "19/05/26-...", _make_days_data())
        assert ws is not None

    def test_replaces_existing_tab(self):
        from helpers.xlsx import write_xlsx_tab
        import openpyxl
        wb = openpyxl.Workbook()
        write_xlsx_tab(wb, "19/05/26-...", _make_days_data())
        write_xlsx_tab(wb, "19/05/26-...", _make_days_data())
        # Should still have exactly one sheet with this name
        assert wb.sheetnames.count("19-05-26-...") == 1

    def test_day_headers_written(self):
        from helpers.xlsx import write_xlsx_tab
        import openpyxl
        wb = openpyxl.Workbook()
        ws = write_xlsx_tab(wb, "19/05/26-...", _make_days_data())
        # Collect all cell values
        all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Dia 1" in all_values
        assert "Dia 2" in all_values

    def test_exercise_names_written(self):
        from helpers.xlsx import write_xlsx_tab
        import openpyxl
        wb = openpyxl.Workbook()
        ws = write_xlsx_tab(wb, "19/05/26-...", _make_days_data())
        all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Sentadilla" in all_values
        assert "Press plano" in all_values
        assert "Dominada" in all_values

    def test_rep_peso_labels_written(self):
        from helpers.xlsx import write_xlsx_tab
        import openpyxl
        wb = openpyxl.Workbook()
        ws = write_xlsx_tab(wb, "19/05/26-...", _make_days_data())
        all_values = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    all_values.append(v)
        assert "Rep." in all_values
        assert "Peso" in all_values
