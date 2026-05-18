"""
helpers/xlsx.py — Escritura de datos al archivo .xlsx local.

Responsabilidades:
    - Crear o reemplazar un tab en el workbook con los datos de entrenamiento
"""

import openpyxl
from helpers.exercise import exercise_display_name, day_exercise_layout


def write_xlsx_tab(wb, tab_name, days_data):
    """
    Crea o reemplaza un tab en el workbook .xlsx con los datos de entrenamiento.
    Estructura: encabezado de día → fila de series → fila Rep./Peso → filas de ejercicios.

    Parámetros:
        wb        — workbook de openpyxl (el archivo .xlsx abierto)
        tab_name  — nombre del tab a crear (ej: "25-05-26-...")
        days_data — dict {día: [ejercicios]} tal como lo retorna parse_pdf
    """
    # Los nombres de hojas en XLSX no pueden contener "/"
    xlsx_tab_name = tab_name.replace("/", "-")
    if xlsx_tab_name in wb.sheetnames:
        del wb[xlsx_tab_name]   # borramos si ya existe
    ws = wb.create_sheet(xlsx_tab_name, 0)   # 0 = insertamos al frente

    row = 1
    for pos, day_num in enumerate(days_data.keys(), start=1):
        exercises = days_data[day_num]
        if not exercises:
            continue

        # Fila "Dia N" — usamos posición (pos) no el número del PDF
        ws.cell(row=row, column=1, value=f"Dia {pos}")
        row += 1

        # Fila de números de serie: [None, 1, None, 2, None, 3, None, ...] (4 semanas × 3 series)
        series_row = [None]
        for _week in range(4):
            for s in range(1, 4):
                series_row.extend([s, None])   # s = número de serie, None = col Peso vacía
        for col_idx, val in enumerate(series_row, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        row += 1

        # Fila de etiquetas Rep./Peso
        label_row = [None]
        for _week in range(4):
            for _s in range(3):
                label_row.extend(["Rep.", "Peso"])
        for col_idx, val in enumerate(label_row, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        row += 1

        # Una fila por ejercicio
        for row_type, ex in day_exercise_layout(exercises):
            if row_type == "blank":
                row += 1
                continue
            ex_row = [exercise_display_name(ex)]
            week_reps = ex.get("week_reps", [None, None, None, None])
            for week_idx in range(4):
                reps = week_reps[week_idx] if week_reps[week_idx] is not None else [None, None, None]
                for s in range(3):
                    ex_row.append(reps[s] if s < len(reps) else None)
                    ex_row.append(None)   # columna Peso (se llena a mano)
            for col_idx, val in enumerate(ex_row, start=1):
                ws.cell(row=row, column=col_idx, value=val)
            row += 1

        row += 2   # dos filas en blanco entre días

    return ws
